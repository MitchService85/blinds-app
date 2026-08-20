"""Regenerate seed + exporter fixtures from the three delivered xlsx files.

Run from repo root: python3 fixtures/generate_seed.py
Sources live in ~/Downloads (the files actually sent to the factory).
"""
import json
import re

import openpyxl

DL = "/Users/mitchservice/Downloads"


def sixteenths(x):
    return round(float(x) * 16)


def parse_tag(raw):
    """'401-LR1' / '2117 - BR' / '1810 -LR' -> (unit, base, index)"""
    unit, tag = re.split(r"\s*-\s*", raw.strip(), maxsplit=1)
    m = re.match(r"([A-Za-z]+)(\d*)$", tag.strip())
    return unit, m.group(1), int(m.group(2)) if m.group(2) else 0


def arbour_floor(path, label, defaults, deduct_source):
    """deduct_source: 'J' (L4 current format) or 'K-text' (L2 legacy notes)."""
    ws = openpyxl.load_workbook(path)["Window Shades"]
    units = {}
    order = []
    for r in range(10, 300):
        raw = ws[f"A{r}"].value
        if raw is None:
            break
        unit, base, idx = parse_tag(raw)
        if unit == "431":
            # Claude's keypad-test unit (Aug 13) that leaked into the
            # Downloads copy via an app export — never part of the real order.
            continue
        note = ws[f"K{r}"].value or ""
        if deduct_source == "J":
            deduct = ws[f"J{r}"].value
            # strip the standard floor note, keep one-off remarks
            extra = note.replace("TIGHT MEASURES. DRILL HOLES IN FASCIA", "").strip(" .")
        else:
            low = note.lower()
            deduct = ("D" if "both" in low else "Dl" if "left" in low
                      else "Dr" if "right" in low else None) if "deduct" in low else None
            extra = "" if ("deduct" in low or low.strip() == "longer chain") else note
        longer_chain = "longer chain" in note.lower()
        ctrl = ws[f"I{r}"].value
        w = {
            "tag_base": base, "tag_index": idx,
            "widths": [sixteenths(ws[f"E{r}"].value)],
            "height": sixteenths(ws[f"F{r}"].value),
            "control_override": "L" if ctrl == "L" else None,
            "deduct": deduct, "longer_chain": longer_chain, "note": extra,
        }
        if unit not in units:
            units[unit] = []
            order.append(unit)
        units[unit].append(w)
    return {
        "label": label, "defaults": defaults,
        "units": [{"number": u, "status": "done", "windows": units[u]} for u in order],
    }


def charles_floor(path):
    ws = openpyxl.load_workbook(path)["Window Shades"]
    units = {}
    order = []
    prev_key = None
    for r in range(10, 300):
        raw = ws[f"A{r}"].value
        if raw is None:
            break
        unit, base, idx = parse_tag(raw)
        key = (unit, base, idx)
        sx = sixteenths(ws[f"E{r}"].value)
        if key == prev_key:
            units[unit][-1]["widths"].append(sx)
        else:
            w = {
                "tag_base": base, "tag_index": idx,
                "widths": [sx],
                "height": sixteenths(ws[f"F{r}"].value),
                "control_override": None, "deduct": None,
                "longer_chain": False, "note": "",
            }
            if unit not in units:
                units[unit] = []
                order.append(unit)
            units[unit].append(w)
        prev_key = key
    return {
        "label": "Batch 3",
        "defaults": {"roll": False, "drive": "R", "tight": True, "extra_note": "",
                     "d_value": "0.5",
                     "color_codes": {"mbed": "", "liv": "PWS1WHIT",
                                     "bed": "YUNOWH", "kit": ""}},
        "units": [{"number": u, "status": "done", "windows": units[u]} for u in order],
    }


def alcon_floor(path):
    """Office format: no room tags; units are zone runs ("Level 1", "L1- Snake
    Corridor"), one untagged 1-panel window per row, K notes kept per window."""
    ws = openpyxl.load_workbook(path)["Window Shades"]
    units = {}
    order = []
    for r in range(10, 500):
        raw = ws[f"A{r}"].value
        if raw is None:
            break
        zone = raw.strip()
        w = {
            "tag_base": "", "tag_index": 0,
            "widths": [sixteenths(ws[f"E{r}"].value)],
            "height": sixteenths(ws[f"F{r}"].value),
            "control_override": None, "deduct": None,
            "longer_chain": False, "note": ws[f"K{r}"].value or "",
        }
        if zone not in units:
            units[zone] = []
            order.append(zone)
        units[zone].append(w)
    return {
        "label": "All Areas",
        "defaults": {"roll": False, "drive": "R", "tight": False, "extra_note": "",
                     "d_value": "0.5",
                     "color_codes": {"mbed": "", "liv": "PWS1WHIT",
                                     "bed": "YUNOWH", "kit": ""}},
        "units": [{"number": z, "status": "done", "windows": units[z]} for z in order],
    }


def cleveland_floor(path):
    """Office format with the Q column: one row per SIZE, quantity in B
    (13 identical blinds -> one window with quantity 13). Units are levels
    (L12/L11/L10); occasional off-template M-column note kept per window."""
    ws = openpyxl.load_workbook(path)["Window Shades"]
    units = {}
    order = []
    for r in range(10, 200):
        raw = ws[f"A{r}"].value
        if raw is None:
            break
        level = raw.strip()
        note = ws[f"M{r}"].value or ""
        w = {
            "tag_base": "", "tag_index": 0,
            "widths": [sixteenths(ws[f"E{r}"].value)],
            "height": sixteenths(ws[f"F{r}"].value),
            "quantity": int(ws[f"B{r}"].value or 1),
            "control_override": None, "deduct": None,
            "longer_chain": False, "note": note,
        }
        if level not in units:
            units[level] = []
            order.append(level)
        units[level].append(w)
    return {
        "label": "All Levels",
        "defaults": {"roll": False, "drive": "R", "tight": True, "extra_note": "",
                     "d_value": "0.5",
                     "color_codes": {"mbed": "", "liv": "PWS1WHIT",
                                     "bed": "YUNOWH", "kit": ""}},
        "units": [{"number": u, "status": "done", "windows": units[u]} for u in order],
    }


ARBOUR_CODES = {"bed": "YUNOWH", "liv": "PWS1WHIT", "studio": "PWS3PLAT", "kitchen": ""}

seed = {
    "projects": [
        {
            "name": "Arbour House 15 Neighborhood Lane",
            "address": "15 Neighborhood Lane",
            "building_type": "residential",
            "tag_chips": ["LR", "BR", "MBR", "Kit", "Studio", "Den", "Bath"],
            "floors": [
                arbour_floor(
                    f"{DL}/Arbour House 15 Neighborhood Lane - Level 2.xlsx", "Level 2",
                    {"roll": False, "drive": "R", "tight": True, "extra_note": "",
                     "d_value": "0.5", "color_codes": ARBOUR_CODES},
                    "K-text"),
                arbour_floor(
                    f"{DL}/Arbour House 15 Neighborhood Lane - Level 4.xlsx", "Level 4",
                    {"roll": True, "drive": "R", "tight": True,
                     "extra_note": "DRILL HOLES IN FASCIA",
                     "d_value": "0.5", "color_codes": ARBOUR_CODES},
                    "J"),
            ],
        },
        {
            "name": "Alcon 2665 Meadowpine",
            "address": "2665 Meadowpine Blvd",
            "building_type": "commercial",
            "tag_chips": ["Office", "Boardroom", "Reception", "Kitchen", "Corridor"],
            "floors": [alcon_floor(f"{DL}/Alcon 2665 Meadowpine.xlsx")],
        },
        {
            "name": "Cleaveland Clinic - 105 Adelaide St W",
            "address": "105 Adelaide St W",
            "building_type": "commercial",
            "tag_chips": ["Office", "Boardroom", "Reception", "Kitchen", "Corridor"],
            "floors": [cleveland_floor(f"{DL}/Cleaveland Clinic - 105 Adelaide St W.xlsx")],
        },
        {
            "name": "44 Charles Batch 3",
            "address": "44 Charles St",
            "building_type": "residential",
            "tag_chips": ["LR", "BR", "MBR", "K"],
            "floors": [charles_floor(f"{DL}/44 Charles Batch 3.xlsx")],
        },
    ]
}

with open("fixtures/seed-projects.json", "w") as f:
    json.dump(seed, f, indent=1)

# refresh exporter input fixture to the widths-array shape
l4 = seed["projects"][0]["floors"][1]
fixture = {
    "project_name": "Arbour House 15 Neighborhood Lane",
    "floor_label": "Level 4",
    "export_date": "2026-08-12",
    "defaults": l4["defaults"],
    "units": l4["units"],
}
with open("fixtures/level4-input.json", "w") as f:
    json.dump(fixture, f, indent=1)

for p in seed["projects"]:
    for fl in p["floors"]:
        nw = sum(len(u["windows"]) for u in fl["units"])
        npanels = sum(len(w["widths"]) for u in fl["units"] for w in u["windows"])
        print(f"{p['name']} / {fl['label']}: {len(fl['units'])} units, {nw} windows, {npanels} panels")
